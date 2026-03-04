#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from typing import Dict, Tuple, Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

TARGET_PHASES = {"2", "3", "4", "5", "6", "7"}


def load_qa(log_path: Path) -> Dict[Tuple[str, str], Dict[str, Any]]:
    out: Dict[Tuple[str, str], Dict[str, Any]] = {}
    with log_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row = json.loads(line)
            if row.get("event") != "step":
                continue

            phase = str(row.get("phase", ""))
            step_id = str(row.get("step_id", ""))
            if phase not in TARGET_PHASES or not step_id:
                continue

            req = row.get("request") or {}
            resp = row.get("response") or {}
            question = req.get("question")
            answer = resp.get("answer")
            if not question or answer is None:
                continue

            out[(phase, step_id)] = {
                "phase": phase,
                "step_id": step_id,
                "question": question,
                "expected": req.get("expected", ""),
                "answer": answer,
                "score": row.get("score"),
                "max_score": row.get("max_score"),
                "notes": row.get("notes", ""),
            }
    return out


def sort_key(key: Tuple[str, str]) -> Tuple[int, str]:
    return int(key[0]), key[1]


def build_rows(
    openai_data: Dict[Tuple[str, str], Dict[str, Any]],
    ollama_data: Dict[Tuple[str, str], Dict[str, Any]],
) -> List[List[Any]]:
    keys = sorted(set(openai_data.keys()) | set(ollama_data.keys()), key=sort_key)
    rows: List[List[Any]] = []
    for key in keys:
        o = openai_data.get(key)
        l = ollama_data.get(key)
        src = o or l

        phase = src.get("phase", "") if src else ""
        step_id = src.get("step_id", "") if src else ""
        question = src.get("question", "") if src else ""
        expected = src.get("expected", "") if src else ""

        openai_answer = o.get("answer", "") if o else ""
        openai_score = f"{o.get('score')}/{o.get('max_score')}" if o and o.get("score") is not None else ""
        openai_notes = o.get("notes", "") if o else ""

        ollama_answer = l.get("answer", "") if l else ""
        ollama_score = f"{l.get('score')}/{l.get('max_score')}" if l and l.get("score") is not None else ""
        ollama_notes = l.get("notes", "") if l else ""

        rows.append([
            phase,
            step_id,
            question,
            expected,
            openai_answer,
            openai_score,
            openai_notes,
            ollama_answer,
            ollama_score,
            ollama_notes,
        ])
    return rows


def export_excel(rows: List[List[Any]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "OpenAI_vs_Ollama"

    headers = [
        "Phase",
        "Step",
        "Question",
        "Expected",
        "OpenAI Answer",
        "OpenAI Score",
        "OpenAI Notes",
        "Ollama Answer",
        "Ollama Score",
        "Ollama Notes",
    ]
    ws.append(headers)

    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    wrap_cols = {3, 4, 5, 7, 8, 10}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    widths = {
        1: 8,
        2: 12,
        3: 46,
        4: 22,
        5: 70,
        6: 14,
        7: 30,
        8: 70,
        9: 14,
        10: 30,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export OpenAI vs Ollama QA comparison to Excel")
    parser.add_argument(
        "--openai-run-dir",
        default="test-data/test-results/local_v3_verify_rerun_20260304",
        help="OpenAI run directory",
    )
    parser.add_argument(
        "--ollama-run-dir",
        default="test-data/test-results/local_v3_verify_ollama_gemma3_27b_20260304_rerun2",
        help="Ollama run directory",
    )
    parser.add_argument(
        "--output",
        default="test-data/test-results/compare_openai_vs_ollama_latest.xlsx",
        help="Output xlsx file path",
    )
    args = parser.parse_args()

    openai_log = Path(args.openai_run_dir) / "test_log.jsonl"
    ollama_log = Path(args.ollama_run_dir) / "test_log.jsonl"

    if not openai_log.exists():
        raise FileNotFoundError(f"OpenAI log not found: {openai_log}")
    if not ollama_log.exists():
        raise FileNotFoundError(f"Ollama log not found: {ollama_log}")

    openai_data = load_qa(openai_log)
    ollama_data = load_qa(ollama_log)

    rows = build_rows(openai_data, ollama_data)
    out_path = Path(args.output)
    export_excel(rows, out_path)

    print(f"✅ Excel generated: {out_path}")
    print(f"   Rows: {len(rows)}")


if __name__ == "__main__":
    main()
